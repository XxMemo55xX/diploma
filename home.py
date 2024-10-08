import os
import time

from flask import Flask, Blueprint, render_template, request, send_file
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

home = Blueprint('home', __name__)


@home.route('/')
def index():
    return render_template('home.html')


def download(filepath):
    print(filepath)
    return send_file(filepath, as_attachment=True)


@home.route('/download_template')
def download_template():
    os.chdir('C:/DEV/Project/diploma')
    path = os.path.realpath('files/template_files/Template.xlsx')
    return download(path)


@home.route('/download_file')
def download_file():
    os.chdir('C:/DEV/Project/diploma')
    path = os.path.realpath('files/download/S_DataValues.xlsx')
    return download(path)


@home.route('/upload_files', methods=["GET", "POST"])
def upload():

    os.chdir('C:/DEV/Project/diploma')
    driver_path = os.path.realpath("webdrivers/chromedriver.exe")
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument("headless")
    chrome_options.add_argument("disable-gpu")

    path = os.path.realpath('files/upload')
    path_download = os.path.realpath('files/download')
    check = ''
    if request.method == "POST":
        if request.files:
            if len(os.listdir(path)) != 0:
                for filename in os.listdir(path):
                    filepath = os.path.join(path, filename)
                    os.remove(filepath)
            if len(os.listdir(path_download)) != 0:
                for filename in os.listdir(path_download):
                    filepath = os.path.join(path_download, filename)
                    os.remove(filepath)

            upload_file = request.files["excelfile"]
            upload_filename = upload_file.filename
            delay = request.form.get("delay")

            if delay == "" or delay == " ":
                delay = 0.5
            float(delay)

            if upload_filename != "":
                message = ""

                upload_file.save(os.path.join(path, upload_filename))
                data_filename = upload_filename
                data_file_path = os.path.join(path, data_filename)
                download_file_path = os.path.join(path_download, "S_DataValues.xlsx")

                data_tab_headers = pd.read_excel(data_file_path, 'Data', header=None)
                data_tab_value = pd.read_excel(data_file_path, 'Data')
                reader = pd.read_excel(data_file_path, 'Data', header=None, usecols="A:G")
                data_col_count = len(reader.columns)

                wb = load_workbook(filename=data_file_path)
                ws = wb['Data']

                if data_col_count == 4:
                    name = (data_tab_headers.iat[0, 0])
                    address = (data_tab_headers.iat[0, 1])
                    city = (data_tab_headers.iat[0, 2])
                    zip_code = (data_tab_headers.iat[0, 3])

                    if name == "Name" and address == "Address" and city == "City" and zip_code == "Zip-code":

                        ws['F1'] = "S_Name"
                        ws['G1'] = "S_Address"
                        ws['H1'] = "S_City"
                        ws['I1'] = "S_Zip-code"

                        driver_chrome = webdriver.Chrome(executable_path=driver_path, options=chrome_options)
                        driver_chrome.get("https://www.google.com/maps/")

                        for record in range(len(data_tab_value.index)):
                            name_value = (data_tab_value.iat[record, 0])
                            address_value = (data_tab_value.iat[record, 1])
                            city_value = (data_tab_value.iat[record, 2])
                            zip_code_value = (data_tab_value.iat[record, 3])

                            s_address = ''
                            s_city = ''
                            s_zip_code = ''
                            s_name = ''
# accept the cookies - start
                            if record == 0:
                                cookies(driver_chrome)
# accept the cookies - end

                            search = driver_chrome.find_element(by=By.NAME, value='q')

                            if pd.isna(name_value):
                                if address_value and city_value:
                                    s_name = name_f_address_city(driver_chrome, search, delay, address_value, city_value)
                                    if pd.isna(zip_code_value):
                                        if s_name != "N/A":
                                            s_zip_code = zipcode_f_name(driver_chrome, search, delay, s_name)
                                        else:
                                            s_zip_code = zipcode_f_address(driver_chrome, search, delay, address_value)

                                elif address_value and zip_code:
                                    s_city = city_f_zipcode(driver_chrome, search, delay, zip_code_value)
                                    if s_city != "N/A":
                                        s_name = name_f_address_city(driver_chrome, search, delay, address_value, s_city)

                            if pd.isna(address_value):
                                if name_value and (s_address == '' or s_address == "N/A"):
                                    s_address = address_f_name(driver_chrome, search, delay, name_value)
                                    if pd.isna(city_value) and s_address != "N/A" and (s_city == '' or s_city == "N/A"):
                                        s_city = city_f_address(driver_chrome, search, delay, s_address)
                                    elif pd.isna(city_value) and (s_city == '' or s_city == "N/A"):
                                        s_city = city_f_name(driver_chrome, search, delay, name_value)
                                    if pd.isna(zip_code_value) and s_address != "N/A" and (s_zip_code == '' or s_zip_code == "N/A"):
                                        s_zip_code = zipcode_f_address(driver_chrome, search, delay, s_address)
                                    elif pd.isna(zip_code_value) and (s_zip_code == '' or s_zip_code == "N/A"):
                                        s_zip_code = zipcode_f_name(driver_chrome, search, delay, name_value)

                            if pd.isna(city_value):

                                if name_value and (s_city == '' or s_city == "N/A"):
                                    s_city = city_f_name(driver_chrome, search, delay, name_value)
                                    if pd.isna(address_value) and (s_address == '' or s_address == "N/A"):
                                        s_address = address_f_name(driver_chrome, search, delay, name_value)
                                        if pd.isna(zip_code_value) and s_address != "N/A" and (s_zip_code == '' or s_zip_code == "N/A"):
                                            s_zip_code = zipcode_f_address(driver_chrome, search, delay, s_address)

                                elif address_value and (s_city == '' or s_city == "N/A"):
                                    s_city = city_f_address(driver_chrome, search, delay, address_value)
                                    if pd.isna(name_value) and s_city != "N/A" and (s_name == '' or s_name == "N/A"):
                                        s_name = name_f_address_city(driver_chrome, search, delay, address_value, s_city)
                                    elif zip_code_value and s_city == "N/A" and (s_city == '' or s_city == "N/A"):
                                        s_city = city_f_zipcode(driver_chrome, search, delay, zip_code_value)
                                        if s_name == '' or s_name == "N/A":
                                            s_name = name_f_address_city(driver_chrome, search, delay, address_value, s_city)
                                    if pd.isna(zip_code_value) and s_address != "N/A" and (s_zip_code == '' or s_zip_code == "N/A"):
                                        s_zip_code = zipcode_f_address(driver_chrome, search, delay, address_value)

                                elif zip_code_value and (s_city == '' or s_city == "N/A"):
                                    s_city = city_f_zipcode(driver_chrome, search, delay, zip_code_value)

                            if pd.isna(zip_code_value):
                                if name_value and (s_zip_code == '' or s_zip_code == "N/A"):
                                    s_zip_code = zipcode_f_name(driver_chrome, search, delay, name_value)
                                    if pd.isna(city_value) and s_zip_code != "N/A" and (s_city == '' or s_city == "N/A"):
                                        s_city = city_f_zipcode(driver_chrome, search, delay, s_zip_code)
                                    elif pd.isna(city_value) and pd.isna(address_value) is not True and (s_city == '' or s_city == "N/A"):
                                        s_city = city_f_address(driver_chrome, search, delay, address_value)
                                    if pd.isna(address_value) and (s_address == '' or s_address == "N/A"):
                                        s_address = address_f_name(driver_chrome, search, delay, name_value)

                                elif address_value and (s_zip_code == '' or s_zip_code == "N/A"):
                                    s_zip_code = zipcode_f_address(driver_chrome, search, delay, address_value)
                                    if pd.isna(city_value) and s_zip_code != "N/A" and (s_city == '' or s_city == "N/A"):
                                        s_city = city_f_zipcode(driver_chrome, search, delay, s_zip_code)
                                    elif pd.isna(city_value) and (s_city == '' or s_city == "N/A"):
                                        s_city = city_f_address(driver_chrome, search, delay, address_value)
                                    if pd.isna(name_value) and s_city != "N/A" and (s_name == '' or s_name == "N/A"):
                                        s_name = name_f_address_city(driver_chrome, search, delay, address_value, s_city)

                            if s_name == '':
                                s_name = "N/A"
                            if s_address == '':
                                s_address = "N/A"
                            if s_city == '':
                                s_city = "N/A"
                            if s_zip_code == '':
                                s_zip_code = "N/A"

                            if s_city == "N/A" and (s_address != '' or s_address != "N/A"):
                                s_city = s_address.split(",")[1]
                                s_city = s_city.split()[1]
                            if s_zip_code == "N/A" and (s_address != '' or s_address != "N/A"):
                                s_zip_code = s_address.split(",")[1]
                                s_zip_code = s_zip_code.split()[0]

                            data_to_file(name_value, s_name, address_value, s_address, city_value, s_city, zip_code_value, s_zip_code, ws, record)

                            message = "Success! - file is ready"
                            check = 'pass'

                        driver_chrome.quit()
                        wb.save(download_file_path)
                    else:
                        message = "Please use the official version of the template file available on main page"
                else:
                    message = "Please use the official version of the template file available on main page"
            else:
                message = "please reload page and make sure to choose the data file"
        else:
            message = "please reload page and make sure to choose the data file"
    else:
        message = "upload filed!"
    return render_template('summary.html', message=message, check=check)


def data_to_file(name_value, s_name, address_value, s_address, city_value, s_city, zip_code_value, s_zip_code, ws, record):
    row = record + 2
    if pd.isna(name_value) is not True:
        ws.cell(row, 6, name_value)
    else:
        ws.cell(row, 6, s_name)
    if pd.isna(address_value) is not True:
        ws.cell(row, 7, address_value)
    else:
        ws.cell(row, 7, s_address)
    if pd.isna(city_value) is not True:
        ws.cell(row, 8, city_value)
    else:
        ws.cell(row, 8, s_city)
    if pd.isna(zip_code_value) is not True:
        ws.cell(row, 9, zip_code_value)
    else:
        ws.cell(row, 9, s_zip_code)


@home.route('/home')
def home_return():
    return render_template('home.html')


def cookies(driver_chrome):
    delay_cookie = 10
    try:
        myElem = WebDriverWait(driver_chrome, delay_cookie).until(EC.visibility_of_element_located((By.XPATH, '/html/body/c-wiz/div/div/div/div[2]/div[1]/div[3]/div[1]/div[1]/form[2]/div/div/button')))
        time.sleep(1)
    except TimeoutException:
        print("Loading took too much time!")
    try:
        driver_chrome.find_element(by=By.XPATH, value="/html/body/div[2]/div[2]/div[3]/span/div/div/div/div[3]/button[2]/div").click()
    except Exception:
        pass
    try:
        driver_chrome.find_element(by=By.XPATH, value="/html/body/c-wiz/div/div/div/div[2]/div[1]/div[3]/div[1]/div[1]/form[2]/div/div/button").click()
    except Exception:
        pass


def page_to_load(driver_chrome, delay):
    try:
        myElem = WebDriverWait(driver_chrome, 6).until(EC.visibility_of_element_located((By.CLASS_NAME, 'S9kvJb')))

    except TimeoutException:
        print("Loading took too much time!")
    try:
        myElem = WebDriverWait(driver_chrome, 6).until(EC.visibility_of_element_located((By.CLASS_NAME, 'TIHn2')))

    except TimeoutException:
        print("Loading took too much time!")
    try:
        myElem = WebDriverWait(driver_chrome, 6).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div')))

    except TimeoutException:
        print("Loading took too much time!")
    time.sleep(float(delay))


def address_f_name(driver_chrome, search, delay, name_value):
    try:
        driver_chrome.find_element(by=By.NAME, value='q').clear()
    except Exception:
        pass

    search.send_keys(name_value)
    search.send_keys(Keys.RETURN)

    page_to_load(driver_chrome, delay)
    time.sleep(1)
    try:
        s_address = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[7]/div[1]/button/div[1]/div[2]').text
    except NoSuchElementException:
        s_address="N/A"

    if s_address=="N/A" or s_address=="":
        try:
            s_address = driver_chrome.find_element(by=By.CSS_SELECTOR, value='[data-tooltip="Copy address"]').text
        except NoSuchElementException:
            s_address="N/A"

    if s_address=="N/A" or s_address=="":
        try:
            s_address = driver_chrome.find_element(by=By.CSS_SELECTOR, value='[data-tooltip="Kopiuj adres"]').text
        except NoSuchElementException:
            s_address="N/A"

    if s_address=="N/A" or s_address=="":
        try:
            s_address = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[7]/div[3]/button/div[1]').text
        except NoSuchElementException:
            s_address="N/A"

    if s_address=="N/A" or s_address=="":
        try:
            s_address = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[7]/div[3]/button/div[1]/div[2]/div[1]').text
        except NoSuchElementException:
            s_address="N/A"

    if s_address=="N/A" or s_address=="":
        try:
            s_address = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div[1]/div[1]/h2[1]/span').text
        except NoSuchElementException:
            s_address="N/A"

    if s_address=="N/A" or s_address=="":
        try:
            s_address = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div[1]/div[1]/h2[1]').text
        except NoSuchElementException:
            s_address="N/A"

    if s_address=="N/A" or s_address=="":
        try:
            s_address = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[9]/div[1]/button/div[1]').text
        except NoSuchElementException:
            s_address="N/A"

    if s_address=="":
        s_address="N/A"
    return s_address


def city_f_zipcode(driver_chrome, search, delay, zip_code_value):
    try:
        driver_chrome.find_element(by=By.NAME, value='q').clear()
    except Exception:
        pass

    search.send_keys(zip_code_value)
    search.send_keys(Keys.RETURN)

    page_to_load(driver_chrome, delay)
    time.sleep(1)
    try:
        s_city = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div[1]/div[1]/div[1]/h1/span[1]').text
    except NoSuchElementException:
        s_city="N/A"

    try:
        if s_city != "N/A":
            temp_city = s_city
            s_city = temp_city.split()[1]
    except Exception:
        print("too fast")

    if s_city=="":
        s_city="N/A"
    return s_city


def name_f_address_city(driver_chrome, search, delay, address_value, city_value):
    try:
        driver_chrome.find_element(by=By.NAME, value='q').clear()
    except Exception:
        pass

    search.send_keys(address_value + " " + city_value)
    search.send_keys(Keys.RETURN)

    page_to_load(driver_chrome, delay)
    time.sleep(1)
    try:
        s_name = driver_chrome.find_element(by=By.CLASS_NAME, value="bfdHYd Ppzolf")
        s_name = s_name.get_attribute("aria-label")
    except NoSuchElementException:
        s_name="N/A"

    if s_name=="N/A" or s_name.upper()==address_value.upper():
        try:
            s_name = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div[1]/div[1]/div[1]/h1').text
        except NoSuchElementException:
            s_name="N/A"

    if s_name=="N/A" or s_name.upper()==address_value.upper():
        try:
            s_name = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[15]/div/div[2]/div[2]/div[1]/div/div/div/div[1]/div/span').text
        except NoSuchElementException:
            s_name="N/A"

    if s_name=="N/A" or s_name.upper()==address_value.upper():
        try:
            s_name = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[17]/div/div[2]/div[2]/div[1]/div/div/div/div[1]').text
        except NoSuchElementException:
            s_name="N/A"

    if s_name=="N/A" or s_name.upper()==address_value.upper():
        try:
            s_name = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[17]/div/div[2]/div[2]/div[1]/div/div/div/div[1]').text
        except NoSuchElementException:
            s_name="N/A"

    if s_name=="N/A" or s_name.upper()==address_value.upper():
        try:
            s_name = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[19]/div/div[2]/div[2]/div[1]/div/div/div/div[1]/div').text
        except NoSuchElementException:
            s_name="N/A"

    if s_name=="N/A" or s_name.upper()==address_value.upper():
        try:
            s_name = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[15]/div/div[2]/div[2]/div[1]/div/div/div/div[1]/div').text
        except NoSuchElementException:
            s_name="N/A"

    if s_name=="N/A" or s_name.upper()==address_value.upper():
        try:
            s_name = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[27]/div/div[2]/div[2]/div[1]/div/div/div/div[1]/div').text
        except NoSuchElementException:
            s_name="N/A"

    if s_name=="N/A" or s_name.upper()==address_value.upper():
        try:
            s_name = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[18]/div/div[2]/div[2]/div[1]/div/div/div/div[1]/div').text
        except NoSuchElementException:
            s_name="N/A"

    if s_name=="":
        s_name="N/A"
    return s_name


def city_f_name(driver_chrome, search, delay, name_value):
    try:
        driver_chrome.find_element(by=By.NAME, value='q').clear()
    except Exception:
        pass
    exception_check = 0

    search.send_keys(name_value)
    search.send_keys(Keys.RETURN)

    page_to_load(driver_chrome, delay)
    time.sleep(1)
    try:
        s_city = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div[1]/div[1]/h2[2]/span').text
        exception_check = 1
    except NoSuchElementException:
        s_city="N/A"

    if s_city=="N/A" or s_city=="":
        try:
            s_city = driver_chrome.find_element(by=By.CSS_SELECTOR, value='[data-tooltip="Copy address"]').text
        except NoSuchElementException:
            s_city="N/A"

    if s_city=="N/A" or s_city=="":
        try:
            s_city = driver_chrome.find_element(by=By.CSS_SELECTOR, value='[data-tooltip="Kopiuj adres"]').text
        except NoSuchElementException:
            s_city="N/A"

    if s_city=="N/A" or s_city=="":
        try:
            s_city = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[7]/div[2]/button/div[1]/div[2]/div[1]').text
        except NoSuchElementException:
            s_city="N/A"

    if s_city=="N/A" or s_city=="":
        try:
            s_city = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[11]/div[3]/button/div[1]/div[2]/div[1]').text
        except NoSuchElementException:
            s_city="N/A"

    if s_city=="N/A" or s_city=="":
        try:
            s_city = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[7]/div[3]/button/div[1]/div[2]/div[1]').text
        except NoSuchElementException:
            s_city="N/A"

    if s_city=="N/A" or s_city=="":
        try:
            s_city = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[17]/div[3]/button/div[1]/div[2]/div[1]').text
        except NoSuchElementException:
            s_city="N/A"

    if s_city=="N/A" or s_city=="":
        try:
            s_city = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[9]/div[2]/button/div[1]/div[2]/div[1]').text
        except NoSuchElementException:
            s_city="N/A"

    if s_city=="N/A" or s_city=="":
        try:
            s_city = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[15]/div[3]/button/div[1]/div[2]/div[1]').text
        except NoSuchElementException:
            s_city="N/A"
    try:
        if s_city!="N/A" or s_city!="":
            temp_city = s_city
            if exception_check == 0:
                temp_city = temp_city.split(sep=",")[1]
            s_city = temp_city.split()[1]
    except Exception:
        s_city="N/A"

    if s_city=="":
        s_city="N/A"
    return s_city


def zipcode_f_name(driver_chrome, search, delay, name_value):
    try:
        driver_chrome.find_element(by=By.NAME, value='q').clear()
    except Exception:
        pass
    exception_check = 0

    search.send_keys(name_value)
    search.send_keys(Keys.RETURN)

    page_to_load(driver_chrome, delay)
    time.sleep(1)
    try:
        s_zip_code = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div[1]/div[1]/h2[2]/span').text
        exception_check = 1
    except NoSuchElementException:
        s_zip_code="N/A"

    if s_zip_code=="N/A" or s_zip_code=="":
        try:
            s_zip_code = driver_chrome.find_element(by=By.CSS_SELECTOR, value='[data-tooltip="Copy address"]').text
        except NoSuchElementException:
            s_zip_code="N/A"

    if s_zip_code=="N/A" or s_zip_code=="":
        try:
            s_zip_code = driver_chrome.find_element(by=By.CSS_SELECTOR, value='[data-tooltip="Kopiuj adres"]').text
        except NoSuchElementException:
            s_zip_code="N/A"

    if s_zip_code=="N/A" or s_zip_code=="":
        try:
            s_zip_code = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[7]/div[2]/button/div[1]/div[2]/div[1]').text
        except NoSuchElementException:
            s_zip_code="N/A"

    if s_zip_code=="N/A" or s_zip_code=="":
        try:
            s_zip_code = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[11]/div[3]/button/div[1]/div[2]/div[1]').text
        except NoSuchElementException:
            s_zip_code="N/A"

    if s_zip_code=="N/A" or s_zip_code=="":
        try:
            s_zip_code = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[7]/div[3]/button/div[1]/div[2]/div[1]').text
        except NoSuchElementException:
            s_zip_code="N/A"

    if s_zip_code=="N/A" or s_zip_code=="":
        try:
            s_zip_code = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[17]/div[3]/button/div[1]/div[2]/div[1]').text
        except NoSuchElementException:
            s_zip_code="N/A"

    if s_zip_code=="N/A" or s_zip_code=="":
        try:
            s_zip_code = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[9]/div[2]/button/div[1]/div[2]/div[1]').text
        except NoSuchElementException:
            s_zip_code="N/A"

    if s_zip_code=="N/A" or s_zip_code=="":
        try:
            s_zip_code = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[15]/div[3]/button/div[1]/div[2]/div[1]').text
        except NoSuchElementException:
            s_zip_code="N/A"
    try:
        if s_zip_code!="N/A" or s_zip_code!="":
            temp_zip_code = s_zip_code
            if exception_check == 0:
                temp_zip_code = temp_zip_code.split(sep=",")[1]
            s_zip_code1 = temp_zip_code.split()[0]
            s_zip_code = s_zip_code1
    except Exception:
        s_zip_code="N/A"

    if s_zip_code=="":
        s_zip_code="N/A"
    return s_zip_code


def city_f_address(driver_chrome, search, delay, address_value):
    try:
        driver_chrome.find_element(by=By.NAME, value='q').clear()
    except Exception:
        pass

    search.send_keys(address_value)
    search.send_keys(Keys.RETURN)

    page_to_load(driver_chrome, delay)
    time.sleep(1)
    try:
        s_city = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div[1]/div[1]/div/div[2]/div[2]/div[1]/div/div/div/div[4]/div').text
    except NoSuchElementException:
        s_city="N/A"

    if s_city=="N/A" or s_city=="" or s_city.upper() == address_value.upper():
        try:
            s_city = driver_chrome.find_element(by=By.CSS_SELECTOR, value='[data-tooltip="Copy address"]').text
        except NoSuchElementException:
            s_city="N/A"

    if s_city=="N/A" or s_city=="" or s_city.upper() == address_value.upper():
        try:
            s_city = driver_chrome.find_element(by=By.CSS_SELECTOR, value='[data-tooltip="Kopiuj adres"]').text
        except NoSuchElementException:
            s_city="N/A"

    if s_city=="N/A" or s_city=="" or s_city.upper() == address_value.upper():
        try:
            s_city = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div[1]/div[1]/h2/span').text
        except NoSuchElementException:
            s_city="N/A"

    if s_city=="N/A" or s_city=="" or s_city.upper() == address_value.upper():
        try:
            s_city = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div[1]/div[1]/h2[3]/span').text
        except NoSuchElementException:
            s_city="N/A"

    if s_city=="N/A" or s_city=="" or s_city.upper() == address_value.upper():
        try:
            s_city = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div[1]/div[1]/h2[2]/span').text
        except NoSuchElementException:
            s_city="N/A"

    try:
        if s_city!="N/A" or s_city!="":
            temp_city = s_city
            s_city = temp_city.split(",")[1]
            temp_city = s_city
            s_city = temp_city.split()[1]
    except Exception:
        s_city="N/A"

    if s_city=="":
        s_city="N/A"

    return s_city


def zipcode_f_address(driver_chrome, search, delay, address_value):
    try:
        driver_chrome.find_element(by=By.NAME, value='q').clear()
    except Exception:
        pass

    search.send_keys(address_value)
    search.send_keys(Keys.RETURN)

    page_to_load(driver_chrome, delay)
    time.sleep(1)
    try:
        s_zip_code = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div[1]/div[1]/div/div[2]/div[2]/div[1]/div/div/div/div[4]/div').text
    except NoSuchElementException:
        s_zip_code="N/A"

    if s_zip_code=="N/A" or s_zip_code=="" or s_zip_code.upper() == address_value.upper():
        try:
            s_zip_code = driver_chrome.find_element(by=By.CSS_SELECTOR, value='[data-tooltip="Copy address"]').text
        except NoSuchElementException:
            s_zip_code="N/A"

    if s_zip_code=="N/A" or s_zip_code=="" or s_zip_code.upper() == address_value.upper():
        try:
            s_zip_code = driver_chrome.find_element(by=By.CSS_SELECTOR, value='[data-tooltip="Kopiuj adres"]').text
        except NoSuchElementException:
            s_zip_code="N/A"

    if s_zip_code=="N/A" or s_zip_code=="" or s_zip_code.upper() == address_value.upper():
        try:
            s_zip_code = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div[1]/div[1]/h2/span').text
        except NoSuchElementException:
            s_zip_code="N/A"

    if s_zip_code=="N/A" or s_zip_code=="" or s_zip_code.upper() == address_value.upper():
        try:
            s_zip_code = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div[1]/div[1]/h2[3]/span').text
        except NoSuchElementException:
            s_zip_code="N/A"

    if s_zip_code=="N/A" or s_zip_code=="" or s_zip_code.upper() == address_value.upper():
        try:
            s_zip_code = driver_chrome.find_element(by=By.XPATH, value='//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div[1]/div[1]/h2[2]/span').text
        except NoSuchElementException:
            s_zip_code="N/A"

    try:
        if s_zip_code!="N/A" or s_zip_code!="" or s_zip_code.upper() != address_value.upper():
            temp_zip_code = s_zip_code
            s_zip_code = temp_zip_code.split(",")[1]
            temp_zip_code = s_zip_code
            s_zip_code = temp_zip_code.split()[0]
    except Exception:
        s_zip_code="N/A"

    if s_zip_code == "" or s_zip_code.upper() == address_value.upper():
        s_zip_code="N/A"

    return s_zip_code
